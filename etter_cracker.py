#!/usr/bin/python3

import openpyxl
import re
from sys import argv

'''
Author: Mark M.
Description: Simple script that deobfuscates XLSM/4.0 maldocs
Date: 5-05-2021
'''

class ettercracker:
	#init
	def __init__(self, path):
		self.path = path
		self.wb_obj = openpyxl.load_workbook(path)
		self.deobfuscated_call = ''

	#find all sheet names and initiate deobfuscation
	def deobfuscate(self):
		self.names = self.wb_obj.sheetnames
		for i in self.names:
			sheet=self.wb_obj.get_sheet_by_name(i)
			self.get_content(sheet)
		return self.deobfuscated_call

	#iterate through all sheet cell values and parse interesting cells that contain values
	def get_content(self, sheet):
		max_col = sheet.max_column
		m_row = sheet.max_row
		re_coords = re.compile("[A-Z]{1,4}[0-9]{1,4}")
		for i in range (1, m_row+1):
			for j in range(1, max_col + 1):
				cell_obj = sheet.cell(row = i, column = j)
				has_coord = re.findall(re_coords, str(cell_obj.value))
				if len(has_coord) > 0:
					call = cell_obj.value
					try:
						self.deobf_call(call, sheet.title)
						self.deobfuscated_call += "\n\n"
					except:
						pass
					
	#differentiate individual strings from cell coordinates	
	def deobf_call(self, call, title):
		for j in call.split("&"):
			lines = j.split(",")
			for line in lines:
				try:
					self.parse_cell(line, title)
				except:
					self.deobfuscated_call += line
		
	
	#parse Excel cell coordinates from interesting cells and append values
	def parse_cell(self, line, title):
		re_coords = re.findall(r"^(.+\()?([A-Z]{1,4}[0-9]{1,4})", line)
		if re_coords[0][1]:
			fname = re_coords[0][1]
			coords = fname
			sheet_name  = title
		else:
			fname =re.findall("(((?<=\().+![A-Z]{1,3}[0-9]{1,3})|([a-zA-Z0-9]{3,20}![A-Z]{1,3}[0-9]{1,3}))", line)[0][0]
			coords = fname.split("!")[1]
			sheet_name = fname.split("!")[0]
		try:
			pretext = line.split(fname)[0]
		except:
			pretext = ""
		try:
			atext=line.split(fname)[1]
		except:
			pretext = ""
		col_raw = re.findall("[A-Z]{1,3}", coords)[0]
		row = re.findall("[0-9]{1,3}", coords)[0]
		cellcol = openpyxl.utils.cell.column_index_from_string(col_raw)
		sheet = self.wb_obj.get_sheet_by_name(sheet_name)
		value = str(sheet.cell(row=int(row), column=int(cellcol)).value)

		self.find_recursive(value, title)
		self.deobfuscated_call += pretext + value + atext

	def find_recursive(self, line, title):
		if "concatenate(" in line.lower() or "exec(" in line.lower() or "call(" in line.lower():
			self.deobf_call(line, title)
		else:
			pass
	
#Attempt to extract URLs from command	
def parse_deobfuscated_url(command):
	urls = re.findall("(http://|ftp://|https://)([\w_-]+(?:(?:\.[\w_-]+)+))([\w.,@?^%&:/~+#-]*[\w@?%&/~+#-])", command)
	for url in urls:
		complete_url = ''
		for item in url:
			complete_url += item
		print(f"{bcolors.HEADER}+URL Extracted:{bcolors.ENDC} {complete_url}")

#Pretty lights
class bcolors:
	HEADER = '\033[95m'
	BLUE = '\033[94m'
	CYAN = '\033[96m'
	GREEN = '\033[92m'
	WARNING = '\033[93m'
	RED = '\033[91m'
	ENDC = '\033[0m'
	BOLD = '\033[1m'
	UNDERLINE = '\033[4m'

def main():
	try:
		path = argv[1]
	except:
		path = input("Enter Filepath: ")
	print(f"{bcolors.HEADER}+ Beginning Deobfuscation of {path}{bcolors.ENDC}")	
	deobf_obj = ettercracker(path)
	command = deobf_obj.deobfuscate()
	print(f"{bcolors.HEADER}+ Extracted Function Calls:{bcolors.ENDC} {command}\n")
	urls = parse_deobfuscated_url(command)

if __name__=='__main__':
	main()
